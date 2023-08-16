#############  analyze_quality_sheets.py  ###########
#
# Created by Andrew Gumbert, June 2023 for Thomas Hanson Lab
# Last updated by Andrw Gumbert, August 2023
#
# Processes all excel files in "data_quality_sheets" folder.
# Expects file names to have the form "<Subject ID>_ERP_quality_sheet"
# For example: "IP10_ERP_quality_sheet"
#
# Expects data to be on sheet2
#
# Will compute average of each time reading, as well as the 
# standard deviation, average + 2*STD, and average - 2*STD.
#
# Highlights individual readings beyond two standard deviations 
# from the mean in red.
#
# Saves resulting colored file as "color_" + <original_filename>
# Saves these processed files to "data_quality_sheets_color"
# folder, which is in the same directory as the "data_quality_sheets"
# folder.
#
# INCLUDES AF3 AND AF4 FROM AVERAGE AND STANDARD DEVIATION




# import packages 

import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import math


# for Tom's computer:
path = "S:\PROJECTS\InfoPos\data_quality_sheets"


# for Andrew's computer:
#path = "/Volumes/as_rsch_NCL02$/PROJECTS/InfoPos/data_quality_sheets"

for filename in os.listdir(path):

    # sets up file reading 
    if filename[-23:] == "_ERP_quality_sheet.xlsx":
        wb = load_workbook(path + "/" + filename)
        wb.active = wb['Sheet2'] # goes to sheet 2
        sheet = wb.active

        # establishes fills for data highlighting 
        greenFill = PatternFill("solid", fgColor = "00FF00")
        redFill = PatternFill("solid", fgColor = "FF0000")

        # titles bottom rows that will be calculated 
        sheet['A36'] = 'Average'
        sheet['A37'] = 'SD'
        sheet['A38'] = '2SD over'
        sheet['A39'] = '2SD under'

        # loops through each of 14 columns with data 
        for cl in range(14):

            # computes average and notes it in row 36 
            average = 0
            for rw in range(29):
                average += float(sheet.cell(row=(rw + 4), column=(cl + 2)).value)
            average /= 29
            sheet.cell(row=36, column=(cl + 2)).value = average
            sheet.cell(row=36, column=(cl + 2)).fill = greenFill

            # computes standard deviation 
            std = 0
            for rw in range(29):
                variance = float(sheet.cell(row=(rw + 4), column=(cl + 2)).value)
                variance -= average
                variance = variance * variance
                std += variance
            std /= 29
            std = math.sqrt(std)
            
            # fills rows 37 - 39 with standard dev and limits of normal range 
            sheet.cell(row=37, column=(cl + 2)).value = std 
            sheet.cell(row=37, column=(cl + 2)).fill = greenFill
            sheet.cell(row=38, column=(cl + 2)).value = average + (2 * std)
            sheet.cell(row=38, column=(cl + 2)).fill = greenFill
            sheet.cell(row=39, column=(cl + 2)).value = average - (2 * std)
            sheet.cell(row=39, column=(cl + 2)).fill = greenFill

            # highlights cells in red that are outside normal range 
            for rw in range(34):
                reading = float(sheet.cell(row=(rw + 2), column=(cl + 2)).value)

                if reading > average +  (2 * std) or reading < average - (2 * std):
                    sheet.cell(row=(rw + 2), column=(cl + 2)).fill = redFill


        # saves sheet
        wb.save(path + "_color/color_" + filename)
